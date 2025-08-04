# エクセルファイルやそのシートについて、本プロジェクト特有の処理

from string import ascii_uppercase
from itertools import product
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

header_bgcolor = PatternFill(start_color='AED6F1', fill_type='solid')

def workbook_column_names():
    length = 1
    while True:
        for letters in product(ascii_uppercase, repeat=length):
            yield ''.join(letters)
        length += 1

summary_columns = [
    ('id', 'ID'),
    ('sheet', 'シート'),
    ('subsystem_name', 'サブシステム名'),
    ('target_type', 'ページ/アドオン'),
    ('category', '機能分類'),
    ('scenario_name', 'シナリオ名'),
    ('description', '概要'),
    ('link', 'リンク'),
    ('result', 'テスト結果'),
    ('ticket', '関連チケット'),
    ('owner', '担当'),
    ('date', '実施日'),
    ('comment', 'コメント'),
    ('confirmed', '修正確認'),
    ('confirm_date', '確認日'),
]

def create_result_workbook():
    wb = openpyxl.Workbook()
    summary_sheet = wb.worksheets[0]
    summary_sheet.title = 'サマリ'
    column_width = summary_sheet.column_dimensions['A'].width * 1.25
    for colname, (_, header_text) in zip(workbook_column_names(), summary_columns):
        summary_sheet.column_dimensions[colname].width = column_width
        summary_sheet[f'{colname}1'] = header_text
        summary_sheet[f'{colname}1'].fill = header_bgcolor
    return wb

case_result_sheet_headers = {
    'formal': [
        ('id', 'ID'),
        ('subsystem_name', 'サブシステム名'),
        ('category', '分類'),
        ('target_type', 'ページ/アドオン'),
        ('reserved_1', ''),
        ('reserved_2', ''),
        ('reserved_3', ''),
        ('designer', '作成者'),
        ('create_date', '作成日'),
        ('fix_date', '修正日'),
    ],
    'semantical': [
        ('abstract', '概要'),
        ('reserved_1', ''),
        ('required_data', '用意するテストデータ'),
        ('result', 'テスト結果'),
        ('ticket_url', '関連チケットURL'),
        ('owner', '担当'),
        ('date', '実施日'),
        ('comment', 'コメント'),
        ('confirmed', '修正確認'),
        ('confirm_date', '確認日'),
    ],
    'steps': [
        ('index', 'No.'),
        ('title', 'テスト手順'),
        ('purpose', '確認内容'),
        ('succeeded', '実施'),
        ('comment', 'コメント'),
        ('who_executed', '実施者'),
        ('date', '実施日'),
        ('screenshot', 'スクリーンショット'),
        ('reserved_1', ''),
        ('reserved_2', ''),
    ],
}

def add_case_result_sheet(wb, step_seq_result):
    sheet = wb.create_sheet(step_seq_result['step_seq_id'])
    for colname, (_, header_text) in zip(workbook_column_names(), case_result_sheet_headers['formal']):
        sheet[f'{colname}1'] = header_text
        sheet[f'{colname}1'].fill = header_bgcolor
    sheet['A2'] = step_seq_result['step_seq_id']
    sheet['B2'] = step_seq_result['サブシステム名']
    sheet['C2'] = step_seq_result['機能分類']
    sheet['D2'] = step_seq_result['ページ/アドオン']
    sheet['H2'] = step_seq_result['author']
    sheet['I2'] = step_seq_result['today']
    sheet['J2'] = ''

    for colname, (_, header_text) in zip(workbook_column_names(), case_result_sheet_headers['semantical']):
        sheet[f'{colname}4'] = header_text
        sheet[f'{colname}4'].fill = header_bgcolor
    sheet['A5'] = step_seq_result['概要'] if '概要' in step_seq_result else step_seq_result['title']
    sheet['A5'].alignment = Alignment(wrap_text=True)
    sheet.merge_cells('A4:B4')
    sheet.merge_cells('A5:B5')
    sheet['C5'] = step_seq_result['用意するテストデータ']
    sheet['C5'].alignment = Alignment(wrap_text=True)
    sheet['D5'] = '成功'
    sheet['E5'] = f'GRDM-{step_seq_result["ticket_number"]}'
    sheet['E5'].hyperlink = f'https://redmine.devops.rcos.nii.ac.jp/issues/{step_seq_result["ticket_number"]}'
    sheet['F5'] = step_seq_result['author']
    sheet['G5'] = step_seq_result['today']
    sheet['H5'] = ''
    sheet['I5'] = ''
    sheet['J5'] = ''

    for cell in sheet['A5:J5'][0]:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    sheet['A6'] = '確認環境'
    sheet['A6'].fill = header_bgcolor
    sheet['B6'] = 'Ubuntu'
    sheet['C6'] = 'Chrome(Playwright)'
    sheet['D6'] = 'ja-JP'

    for colname, (_, header_text) in zip(workbook_column_names(), case_result_sheet_headers['steps']):
        sheet[f'{colname}8'] = header_text
        sheet[f'{colname}8'].fill = header_bgcolor

    cell_standard_width = sheet.column_dimensions['A'].width
    sheet.column_dimensions['B'].width = cell_standard_width * 4
    sheet.column_dimensions['C'].width = cell_standard_width * 5
    sheet.column_dimensions['E'].width = cell_standard_width * 2
    sheet.column_dimensions['G'].width = cell_standard_width * 2
    sheet.column_dimensions['H'].width = cell_standard_width * 2
    sheet.column_dimensions['I'].width = cell_standard_width * 2
    sheet.column_dimensions['J'].width = cell_standard_width * 2
    sheet.row_dimensions[5].height = cell_standard_width * 3

    return sheet

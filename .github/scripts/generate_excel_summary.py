#!/usr/bin/env python3
"""
Generate Excel summary from test results.
This script is used by GitHub Actions to create test reports.
"""

import os
import sys
import re
import shutil
import nbformat
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime
from pathlib import Path
from base64 import b64decode


def collect_all_notebooks(result_dir):
    """Recursively collect notebooks with hierarchical sorting."""
    result_path = Path(result_dir)
    notebooks = []
    
    # Get .ipynb files in current directory
    current_notebooks = [
        p for p in result_path.glob('*.ipynb') 
        if '.ipynb_checkpoints' not in str(p)
    ]
    current_notebooks.sort(key=lambda p: p.stat().st_mtime)
    
    for nb in current_notebooks:
        notebooks.append(nb)
        
        # Check for corresponding folder (e.g., "取りまとめ.ipynb" -> "取りまとめ/")
        notebook_folder = result_path / nb.stem
        if notebook_folder.exists() and notebook_folder.is_dir():
            notebooks.extend(collect_all_notebooks(notebook_folder))
    
    # Check other subdirectories without corresponding .ipynb
    for subdir in result_path.iterdir():
        if (subdir.is_dir() and 
            '.ipynb_checkpoints' not in subdir.name and
            not (result_path / f"{subdir.name}.ipynb").exists()):
            notebooks.extend(collect_all_notebooks(subdir))
    
    return notebooks


def has_header1(cell):
    """Check if cell has level 1 header."""
    if cell['cell_type'] != 'markdown':
        return False
    line = cell['source'].split('\n')[0]
    m = re.match(r'^#\s+(.+)', line)
    return m is not None

def has_header2(cell):
    """Check if cell has level 2 header."""
    if cell['cell_type'] != 'markdown':
        return False
    line = cell['source'].split('\n')[0]
    m = re.match(r'^##\s+(.+)', line)
    return m is not None

def parse_cells(notebook_path):
    """Parse notebook cells to extract test sets."""
    with open(notebook_path, 'r', encoding='utf-8') as f:
        nb = nbformat.read(f, as_version=nbformat.NO_CONVERT)
    
    cells = nb['cells']
    test_sets = []
    
    for i, cell in enumerate(cells):
        if has_header1(cell):
            line = cell['source'].split('\n')[0]
            m = re.match(r'^#\s+(.+)', line)
            if m and '報告書出力' not in m.group(1):
                test_sets.append((i, cell))
            elif m and '報告書出力' in m.group(1):
                break
    
    if len(test_sets) == 0 or test_sets[-1][0] != len(cells) - 1:
        test_sets.append((len(cells), None))
    
    return (cells, test_sets)


def save_image(cellindex, image_base64):
    """Save base64 image to file."""
    filename = f'/tmp/screenshot-{cellindex}.png'
    with open(filename, 'wb') as f:
        f.write(b64decode(image_base64))
    return filename

def get_images_from_cell(cellindex, cell):
    """Extract images from cell outputs."""
    if 'outputs' not in cell:
        return None
    images = [out['data']['image/png'] for out in cell['outputs'] if 'data' in out and 'image/png' in out['data']]
    return [save_image(f"{cellindex}-{i}", image) for i, image in enumerate(images)]

def create_workbook(all_test_sets, author, ticket_number, result_dir):
    """Create Excel workbook with test results."""
    wb = openpyxl.Workbook()
    
    # Create summary sheet
    summary_sheet = wb.active
    summary_sheet.title = 'サマリ'
    
    # Setup summary sheet headers
    fill = PatternFill(start_color='AED6F1', fill_type='solid')
    summary_sheet.column_dimensions['A'].width = summary_sheet.column_dimensions['A'].width * 1.25
    headers = ['ID', 'シート', 'サブシステム', 'ページ/アドオン', '機能分類', 'シナリオ名', '概要', 'リンク', 'テスト結果', '関連チケット', '担当', '実施日', 'コメント', '修正確認', '確認日']
    for colname, text in zip('ABCDEFGHIJKLMNO', headers):
        summary_sheet.column_dimensions[colname].width = summary_sheet.column_dimensions['A'].width
        summary_sheet[f'{colname}1'] = text
        summary_sheet[f'{colname}1'].fill = fill
    
    # Process all test sets
    index = 0
    id_prefix = 'T'
    
    for notebook_file, (cells, test_sets) in all_test_sets:
        print(f"Processing {notebook_file}...")
        sheetname = '_'.join(os.path.splitext(os.path.split(notebook_file)[-1])[0].split('-')[1:][::-1][:2])
        
        for i, ((start, header), (end, _)) in enumerate(zip(test_sets, test_sets[1:])):
            index += 1
            test_id = f'{id_prefix}{index:03d}{sheetname}'
            os.makedirs(os.path.join(result_dir, 'screenshots', test_id), exist_ok=True)
            
            # Extract test attributes
            line = header['source'].split('\n')[0]
            m = re.match(r'#\s+(.+)', line)
            title = m.group(1) if m else ''
            attrs = {}
            for line in header['source'].split('\n'):
                m = re.match(r'-\s+([^:]+):\s*(.+)', line)
                if m:
                    attrs[m.group(1)] = m.group(2)
            
            sheet = wb.create_sheet(test_id)
            
            # Apply header styles
            for colname in 'ABCDEFGHIJ':
                sheet[f'{colname}1'].fill = fill
                sheet[f'{colname}4'].fill = fill
            sheet['A6'].fill = fill
            
            sheet['A1'] = 'ID'
            sheet['A2'] = test_id
            sheet['B1'] = 'サブシステム名'
            sheet['B2'] = attrs.get('サブシステム名', '')
            sheet['C1'] = '分類'
            sheet['C2'] = attrs.get('機能分類', '')
            sheet['D2'] = attrs.get('ページ/アドオン', '')
            sheet['H1'] = '作成者'
            sheet['H2'] = author
            sheet['I1'] = '作成日'
            sheet['I2'] = datetime.now().strftime('%Y-%m-%d')
            sheet['J1'] = '修正日'
            sheet['J2'] = ''
        
            sheet['A4'] = '概要'
            sheet['A5'] = attrs['概要'] if '概要' in attrs else title
            sheet['A5'].alignment = Alignment(wrap_text=True)
            sheet.merge_cells('A4:B4')
            sheet.merge_cells('A5:B5')
            sheet['C4'] = '用意するテストデータ'
            sheet['C5'] = attrs.get('用意するテストデータ', '')
            sheet['C5'].alignment = Alignment(wrap_text=True)
            sheet['D4'] = 'テスト結果'
            sheet['D5'] = '成功'  # Will be updated later based on test results
            sheet['E4'] = '関連チケットURL'
            sheet['E5'] = ticket_number
            sheet['F4'] = '担当'
            sheet['F5'] = author
            sheet['G4'] = '実施日'
            sheet['G5'] = datetime.now().strftime('%Y-%m-%d')
            sheet['H4'] = 'コメント'
            sheet['H5'] = ''
            sheet['I4'] = '修正確認'
            sheet['I5'] = ''
            sheet['J4'] = '確認日'
            sheet['J5'] = ''
            for cell in sheet['A5:J5'][0]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
            sheet['A6'] = '確認環境'
            sheet['B6'] = 'Ubuntu'
            sheet['C6'] = 'Chrome(Playwright)'
            sheet['D6'] = 'ja-JP'
        
            sheet.column_dimensions['B'].width = sheet.column_dimensions['A'].width * 4
            sheet.column_dimensions['C'].width = sheet.column_dimensions['A'].width * 5
            sheet.column_dimensions['E'].width = sheet.column_dimensions['A'].width * 2
            sheet.column_dimensions['G'].width = sheet.column_dimensions['A'].width * 2
            sheet.column_dimensions['H'].width = sheet.column_dimensions['A'].width * 2
            sheet.column_dimensions['I'].width = sheet.column_dimensions['A'].width * 2
            sheet.column_dimensions['J'].width = sheet.column_dimensions['A'].width * 2
            sheet.row_dimensions[5].height = sheet.column_dimensions['A'].width * 3
        
            startrow = 8
            sheet[f'A{startrow}'] = 'No.'
            sheet[f'B{startrow}'] = 'テスト手順'
            sheet[f'C{startrow}'] = '確認内容'
            sheet[f'D{startrow}'] = '実施'
            sheet[f'E{startrow}'] = 'コメント'
            sheet[f'F{startrow}'] = '実施者'
            sheet[f'G{startrow}'] = '実施日'
            sheet[f'H{startrow}'] = 'スクリーンショット'
            for colname in 'ABCDEFGHIJ':
                sheet[f'{colname}{startrow}'].fill = fill
        
            itemheight = sheet.column_dimensions['A'].width * 12 #* 6
            itemindex = 1
            last_images = None
            row = startrow
            has_error = False
        
            for i in range(end - (start + 1)):
                cell = cells[i + start + 1]
                if not has_header2(cell):
                    images = get_images_from_cell(i + start + 1, cell)
                    if images:
                        last_images = images
                    continue
                if last_images is not None and len(last_images) > 0:
                    screenshot = openpyxl.drawing.image.Image(last_images[0])
                    screenshot.height = itemheight
                    screenshot.width = int(itemheight / 1080 * 1920)
                    shutil.copy(last_images[0], os.path.join(result_dir, 'screenshots', test_id, '{0:05d}.png'.format(itemindex - 1)))
                # 成功したか？
                output_types = []
                outputs = []
                for next_cell in cells[i + start + 1 + 1:]:
                    if has_header2(next_cell):
                        break
                    if 'outputs' not in next_cell:
                        continue
                    #assert all([o['output_type'] != 'error' for o in next_cell['outputs']]), next_cell['outputs']
                    output_types += [o['output_type'] for o in next_cell['outputs']]
                    outputs += list(next_cell['outputs'])
                output_types = set(output_types)
                if 'error' in output_types or len(output_types) == 0:
                    has_error = True
                line = cell['source'].split('\n')[0]
                m = re.match(r'##\s+(.+)', line)
                row = startrow + itemindex
                sheet[f'A{row}'] = str(itemindex)
                sheet[f'B{row}'] = m.group(1)
                sheet[f'C{row}'] = '\n'.join(cell['source'].split('\n')[1:]).strip()
                sheet[f'D{row}'] = '■' if 'error' not in output_types and len(output_types) > 0 else '□'
                sheet[f'E{row}'] = '' if 'error' not in output_types else '\n'.join([o['evalue'] if 'evalue' in o else o['ename'] for o in outputs if o['output_type'] == 'error'])
                sheet[f'F{row}'] = 'Playwright'
                sheet[f'G{row}'] = datetime.now().strftime('%Y-%m-%d')
                sheet[f'H{row}'] = ''
                for cell in sheet[f'A{row}:H{row}'][0]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                sheet[f'D{row}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                
                sheet.row_dimensions[row].height = itemheight
        
                itemindex += 1
                last_images = None
                
            if last_images is not None and len(last_images) > 0:
                screenshot = openpyxl.drawing.image.Image(last_images[0])
                screenshot.height = itemheight
                screenshot.width = int(itemheight / 1080 * 1920)
                shutil.copy(last_images[0], os.path.join(result_dir, 'screenshots', test_id, '{0:05d}.png'.format(itemindex - 1)))            

            sheet['D5'] = '失敗' if has_error else '成功'

            summaryrow = index + 1
            summary_sheet[f'A{summaryrow}'] = test_id
            summary_sheet[f'B{summaryrow}'] = test_id
            summary_sheet[f'C{summaryrow}'] = attrs['サブシステム名']
            summary_sheet[f'D{summaryrow}'] = attrs['ページ/アドオン']
            summary_sheet[f'E{summaryrow}'] = attrs['機能分類']
            summary_sheet[f'F{summaryrow}'] = attrs['シナリオ名']
            summary_sheet[f'G{summaryrow}'] = title
            summary_sheet[f'H{summaryrow}'] = f'参照: {test_id}'
            summary_sheet[f'H{summaryrow}'].hyperlink = f'#{test_id}!A1'
            summary_sheet[f'I{summaryrow}'] = '成功' if not has_error else '失敗'
            summary_sheet[f'J{summaryrow}'] = ticket_number
            summary_sheet[f'K{summaryrow}'] = author
            summary_sheet[f'L{summaryrow}'] = datetime.now().strftime('%Y-%m-%d')
            for cell in summary_sheet[f'A{summaryrow}:O{summaryrow}'][0]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    return wb


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python generate_excel_summary.py <result_dir> [author] [ticket]")
        sys.exit(1)
    
    result_dir = sys.argv[1]
    author = sys.argv[2] if len(sys.argv) > 2 else 'GitHub Actions'
    ticket_number = sys.argv[3] if len(sys.argv) > 3 else '00000'
    
    # Generate output filename
    date_str = datetime.now().strftime('%Y-%m-%d')
    output_file = Path(result_dir) / f'test-summary-{date_str}.xlsx'
    
    print(f"Collecting notebooks from {result_dir}...")
    notebooks = collect_all_notebooks(result_dir)
    print(f"Found {len(notebooks)} notebooks")
    
    # Parse all notebooks into test sets
    all_test_sets = []
    for notebook_path in notebooks:
        print(f"  - {notebook_path.relative_to(result_dir)}")
        all_test_sets.append((str(notebook_path), parse_cells(str(notebook_path))))
    
    # Generate Excel workbook
    wb = create_workbook(all_test_sets, author, ticket_number, str(result_dir))
    
    # Save workbook
    wb.save(str(output_file))
    print(f"\nExcel summary saved to: {output_file}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())